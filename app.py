from flask import Flask, render_template, request, redirect, session, send_file
from flask import send_from_directory
import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont
import os
import random
from werkzeug.utils import secure_filename
import shutil
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
from io import BytesIO
import time


BASE_DIR = os.path.dirname(os.path.abspath(__file__))

EXCEL_FOLDER = os.path.join(BASE_DIR, "excel_files")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
ID_TEMPLATE_FOLDER = os.path.join(BASE_DIR, "id_templates")

os.makedirs(EXCEL_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

#############################################
# FLASK APP
#############################################

app = Flask(__name__)
app.secret_key = "ruc_secret"

app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

os.makedirs("excel_files", exist_ok=True)


def connect_db():
    return psycopg2.connect(os.environ["DATABASE_URL"])


#############################################
# EXCEL LOCK SYSTEM
#############################################


def wait_for_excel_lock(lock_file, timeout=30):
    start_time = time.time()

    while os.path.exists(lock_file):
        if time.time() - start_time > timeout:
            # remove stale lock
            os.remove(lock_file)
            break
        time.sleep(1)


#############################################
# FIND COLUMN BY HEADER NAME
#############################################


def find_column(sheet, header_name):

    for col in range(1, sheet.max_column + 1):

        value = sheet.cell(row=1, column=col).value

        if value and str(value).strip().upper() == header_name.upper():
            return col

    return None


#############################################
# VALIDATE MASTER TRACKER (GLOBE NLZ)
#############################################


def validate_tracker(old_file, new_file):

    old_wb = load_workbook(old_file)
    new_wb = load_workbook(new_file)

    if "GLOBE NLZ" not in new_wb.sheetnames:
        return "Sheet 'GLOBE NLZ' is missing"

    old_ws = old_wb["GLOBE NLZ"]
    new_ws = new_wb["GLOBE NLZ"]

    #################################
    # CHECK COLUMN HEADERS
    #################################

    for col in range(1, old_ws.max_column + 1):

        old_header = old_ws.cell(row=1, column=col).value
        new_header = new_ws.cell(row=1, column=col).value

        if old_header != new_header:
            return f"Column changed: {old_header}"

    #################################
    # CHECK ROWS WERE NOT DELETED
    #################################

    if new_ws.max_row < old_ws.max_row:
        return "Rows were deleted from the tracker"

    #################################
    # FIND DU ID COLUMN
    #################################

    du_col_old = find_column(old_ws, "DU ID")
    du_col_new = find_column(new_ws, "DU ID")

    if not du_col_old or not du_col_new:
        return "DU ID column missing"

    #################################
    # CHECK DU ID INTEGRITY
    #################################

    old_du_ids = set()
    new_du_ids = set()

    for row in range(2, old_ws.max_row + 1):

        du = old_ws.cell(row=row, column=du_col_old).value

        if du:
            old_du_ids.add(str(du).strip())

    for row in range(2, new_ws.max_row + 1):

        du = new_ws.cell(row=row, column=du_col_new).value

        if du:

            du = str(du).strip()

            if du in new_du_ids:
                return f"Duplicate DU ID detected: {du}"

            new_du_ids.add(du)

    #################################
    # CHECK FOR MISSING DU IDs
    #################################

    missing_du = old_du_ids - new_du_ids

    if missing_du:
        return f"Missing DU IDs detected: {list(missing_du)[:5]}"

    return "OK"


#############################################
# BACKUP FILE
#############################################


def backup_file(file_path, backup_folder):

    if not os.path.exists(backup_folder):
        os.makedirs(backup_folder)

    if os.path.exists(file_path):

        time_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        file_name = os.path.basename(file_path)

        new_name = time_stamp + "_" + file_name

        backup_path = os.path.join(backup_folder, new_name)

        shutil.copy(file_path, backup_path)


#############################################
# MASTER TRACKER
#############################################


@app.route("/master_tracker")
def master_tracker():

    if "admin" not in session:
        return redirect("/")

    return send_file("excel_files/master/NLZ_MASTER_TRACKER.xlsx", as_attachment=True)


#############################################
# UPLOAD MASTER TRACKER
#############################################


@app.route("/upload_master_tracker", methods=["GET", "POST"])
def upload_master_tracker():

    if "admin" not in session:
        return redirect("/")

    master_path = "excel_files/master/NLZ_MASTER_TRACKER.xlsx"

    if request.method == "POST":

        file = request.files["tracker"]

        if file.filename == "":
            return "No file selected"

        upload_path = "excel_files/master/upload_temp.xlsx"
        file.save(upload_path)

        result = validate_tracker(master_path, upload_path)

        if result != "OK":
            os.remove(upload_path)
            return result

        backup_file(master_path, "backups/master")

        os.replace(upload_path, master_path)

        return "Master Tracker Updated Successfully"

    return render_template("upload_master_tracker.html")


#############################################
# LOGIN
#############################################


@app.route("/", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]

        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute(
            """
            SELECT username,password,role
            FROM admins
            WHERE username=%s
            """,
            (username,),
        )

        admin = cursor.fetchone()

        cursor.close()
        conn.close()

        if admin and check_password_hash(admin[1], password):

            session["admin"] = admin[0]
            session["role"] = admin[2]

            return redirect("/dashboard")

        else:
            return "Wrong Username or Password"

    return render_template("login.html")


#############################################
# DASHBOARD WITH STATISTICS
#############################################


@app.route("/dashboard")
def dashboard():

    if "admin" not in session:
        return redirect("/")

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM projects")
    projects = cursor.fetchall()

    cursor.execute("SELECT COUNT(*) FROM projects")
    total_projects = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM employees")
    total_employees = cursor.fetchone()[0]

    cursor.close()
    conn.close()

    total_ids = 0

    if os.path.exists("excel_files"):

        for file in os.listdir("excel_files"):

            path = os.path.join("excel_files", file)

            if os.path.isfile(path) and file.endswith(".xlsx"):

                wb = load_workbook(path)

                if "ID" in wb.sheetnames:

                    ws = wb["ID"]

                    total_ids += ws.max_row - 1

    base_url = request.host_url

    return render_template(
        "dashboard.html",
        projects=projects,
        base_url=base_url,
        total_projects=total_projects,
        total_employees=total_employees,
        total_ids=total_ids,
    )


#############################################
# SEARCH
#############################################


@app.route("/search", methods=["GET", "POST"])
def search():

    conn = connect_db()
    cursor = conn.cursor()

    employees = []

    if request.method == "POST":

        keyword = request.form["search"]

        cursor.execute(
            """
            SELECT e.id,
                   e.first_name,
                   e.last_name,
                   e.position,
                   e.email,
                   e.mobile,
                   p.project_code
            FROM employees e
            JOIN projects p
            ON e.project_id = p.id
            WHERE e.first_name ILIKE %s
               OR e.last_name ILIKE %s
               OR e.email ILIKE %s
            """,
            ("%" + keyword + "%", "%" + keyword + "%", "%" + keyword + "%"),
        )

        employees = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("search.html", employees=employees)


#############################################
# EDIT EMPLOYEE
#############################################


@app.route("/edit_employee/<emp_id>", methods=["GET", "POST"])
def edit_employee(emp_id):

    if "admin" not in session:
        return redirect("/")

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT e.* , p.project_code
        FROM employees e
        JOIN projects p ON e.project_id = p.id
        WHERE e.id=%s
        """,
        (emp_id,),
    )

    emp = cursor.fetchone()

    if request.method == "POST":

        first_name = request.form["first_name"]
        last_name = request.form["last_name"]
        position = request.form["position"]
        email = request.form["email"]
        mobile = request.form["mobile"]

        cursor.execute(
            """
            UPDATE employees
            SET first_name=%s,
                last_name=%s,
                position=%s,
                email=%s,
                mobile=%s
            WHERE id=%s
            """,
            (first_name, last_name, position, email, mobile, emp_id),
        )

        conn.commit()

        cursor.close()
        conn.close()

        return redirect("/search")

    cursor.close()
    conn.close()

    return render_template("edit_employee.html", emp=emp)


#############################################
# CREATE PROJECT + EXCEL TEMPLATE
#############################################


@app.route("/create_project", methods=["GET", "POST"])
def create_project():

    if "admin" not in session:
        return redirect("/")

    if session.get("role") != "admin":
        return "Access Denied"

    if request.method == "POST":

        project_name = request.form["project_name"]
        region = request.form["region"]
        company = request.form["company"]

        project_code = str(random.randint(10000, 99999))

        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute(
            """
            INSERT INTO projects(project_name,region,company,project_code)
            VALUES(%s,%s,%s,%s)
            """,
            (project_name, region, company, project_code),
        )

        conn.commit()

        cursor.close()
        conn.close()

        #################################
        # CREATE EXCEL
        #################################

        # Ensure folder exists

        wb = Workbook()

        yellow = PatternFill(start_color="FFFF00", fill_type="solid")
        bold = Font(bold=True)
        center = Alignment(horizontal="center")

        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws1 = wb.active
        ws1.title = "ACCESS INFO"

        headers = [
            "NAME",
            "COMPANY",
            "DESIGNATION",
            "AREA ASSIGNED",
            "MOBILE NO.",
            "EMAIL",
            "ANDROID OR IPHONE",
            "FTAP IMEI",
            "FTAP EMAIL USED",
            "PHILTOWER IMEI",
            "PHILTOWER EMAIL USED",
        ]

        ws1.append(headers)

        for col in range(1, len(headers) + 1):

            cell = ws1.cell(row=1, column=col)
            cell.fill = yellow
            cell.font = bold
            cell.alignment = center
            cell.border = border

        sheets = [
            "2X2",
            "NBI",
            "CERTIFICATES",
            "eSignature",
            "SEC ID",
            "WAH CERT",
            "ID",
        ]

        for s in sheets:

            ws = wb.create_sheet(s)

            if s == "SEC ID":
                ws.append(["NAME", "SEC NUMBER", "EXPIRY", "IMAGE"])

            elif s == "ID":
                ws.append(["NAME", "ID NUMBER", "EXPIRY", "IMAGE"])

            else:
                ws.append(["NAME", "IMAGE"])

        file_path = "excel_files/" + project_code + ".xlsx"

        wb.save(file_path)

        backup_file(file_path, "backups/excel")

        return redirect("/dashboard")

    return render_template("create_project.html")


#############################################
# EMPLOYEE FORM
#############################################


@app.route("/form/<code>", methods=["GET", "POST"])
def form(code):

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute(
        "SELECT * FROM projects WHERE project_code=%s",
        (code,),
    )

    project = cursor.fetchone()

    if not project:
        cursor.close()
        conn.close()
        return "Invalid Project Link"

    #################################
    # FORM SUBMIT
    #################################

    if request.method == "POST":

        # TEXT DATA
        first_name = request.form["first_name"]
        last_name = request.form["last_name"]
        position = request.form["position"]
        email = request.form["email"]
        mobile = request.form["mobile"]
        phone_type = request.form["phone_type"]
        ftap_imei = request.form["ftap_imei"]
        ftap_email = request.form["ftap_email"]
        philtower_imei = request.form["philtower_imei"]
        philtower_email = request.form["philtower_email"]

        sec_number = request.form["sec_number"]
        sec_expiry = request.form["sec_expiry"]

        full_name = first_name + " " + last_name

        #################################
        # FILES
        #################################

        photo = request.files["photo"]
        nbi = request.files["nbi"]
        certificate = request.files["certificate"]
        signature = request.files["signature"]

        sec_id = request.files.get("sec_id")
        wah_cert = request.files.get("wah_cert")

        #################################
        # LOCAL FILE UPLOADS
        #################################

        os.makedirs("uploads/photos", exist_ok=True)
        os.makedirs("uploads/nbi", exist_ok=True)
        os.makedirs("uploads/certificates", exist_ok=True)
        os.makedirs("uploads/signatures", exist_ok=True)
        os.makedirs("uploads/secid", exist_ok=True)
        os.makedirs("uploads/wah", exist_ok=True)

        # PHOTO
        photo_name = secure_filename(photo.filename)
        photo_path = os.path.join("uploads/photos", photo_name)
        photo.save(photo_path)
        photo_url = photo_path

        # NBI
        nbi_name = secure_filename(nbi.filename)
        nbi_path = os.path.join("uploads/nbi", nbi_name)
        nbi.save(nbi_path)
        nbi_url = nbi_path

        # CERTIFICATE
        cert_name = secure_filename(certificate.filename)
        cert_path = os.path.join("uploads/certificates", cert_name)
        certificate.save(cert_path)
        cert_url = cert_path

        # SIGNATURE
        sign_name = secure_filename(signature.filename)
        sign_path = os.path.join("uploads/signatures", sign_name)
        signature.save(sign_path)
        sign_url = sign_path

        # OPTIONAL FILES
        sec_url = ""
        wah_url = ""

        if sec_id and sec_id.filename != "":
            sec_name = secure_filename(sec_id.filename)
            sec_path = os.path.join("uploads/secid", sec_name)
            sec_id.save(sec_path)
            sec_url = sec_path

        if wah_cert and wah_cert.filename != "":
            wah_name = secure_filename(wah_cert.filename)
            wah_path = os.path.join("uploads/wah", wah_name)
            wah_cert.save(wah_path)
            wah_url = wah_path

        #################################
        # SAVE EMPLOYEE
        #################################

        project_id = project[0]

        cursor.execute(
            """
            INSERT INTO employees(
            project_id,
            first_name,
            last_name,
            position,
            email,
            mobile,
            phone_type,
            ftap_imei,
            ftap_email,
            philtower_imei,
            philtower_email,
            photo
            )
            VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,
            (
                project_id,
                first_name,
                last_name,
                position,
                email,
                mobile,
                phone_type,
                ftap_imei,
                ftap_email,
                philtower_imei,
                philtower_email,
                photo_url,
            ),
        )

        conn.commit()

        #################################
        # OPEN EXCEL
        #################################

        file_path = os.path.join(EXCEL_FOLDER, code + ".xlsx")
        lock_file = file_path + ".lock"

        wait_for_excel_lock(lock_file)

        open(lock_file, "w").close()

        # create file if it doesn't exist
        if not os.path.exists(file_path):
            wb = Workbook()
            wb.save(file_path)

        wb = load_workbook(file_path)

        #################################
        # ACCESS INFO SHEET
        #################################

        ws1 = wb["ACCESS INFO"]

        ws1.append(
            [
                full_name,
                "RUC",
                position,
                project[2],
                mobile,
                email,
                phone_type,
                ftap_imei,
                ftap_email,
                philtower_imei,
                philtower_email,
            ]
        )

        #################################
        # AUTO ADJUST COLUMN WIDTH
        #################################

        for column in ws1.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            ws1.column_dimensions[column_letter].width = max_length + 4

        #################################
        # INSERT IMAGE FUNCTION
        #################################

        def insert_image(sheet, image_path):

            if image_path == "":
                return

            ws = wb[sheet]

            row = ws.max_row + 3

            img = ExcelImage(image_path)

            img.width = 250
            img.height = 250

            ws.row_dimensions[row].height = 210
            ws.row_dimensions[row + 1].height = 25

            ws.column_dimensions["B"].width = 45

            ws.add_image(img, "B" + str(row))

            ws["B" + str(row + 1)] = full_name

        #################################
        # INSERT IMAGES
        #################################

        insert_image("2X2", photo_url)
        insert_image("NBI", nbi_url)
        insert_image("CERTIFICATES", cert_url)
        insert_image("eSignature", sign_url)
        insert_image("WAH CERT", wah_url)

        #################################
        # SEC ID SHEET
        #################################

        ws6 = wb["SEC ID"]

        row = ws6.max_row + 1

        ws6.cell(row=row, column=1).value = full_name
        ws6.cell(row=row, column=2).value = sec_number
        ws6.cell(row=row, column=3).value = sec_expiry

        if sec_url != "":
            img = ExcelImage(sec_url)

            img.width = 200
            img.height = 140

            ws6.row_dimensions[row].height = 110

            ws6.add_image(img, "D" + str(row))

        #################################
        # SAVE EXCEL
        #################################

        print("Updating Excel file:", file_path)
        wb.save(file_path)

        if os.path.exists(lock_file):
            os.remove(lock_file)

        cursor.close()
        conn.close()

        return "Form Submitted Successfully"

    return render_template("form.html", code=code)


#############################################
# OPEN PROJECT EXCEL
#############################################


@app.route("/open_excel/<code>")
def open_excel(code):

    file_path = os.path.join(EXCEL_FOLDER, code + ".xlsx")

    if os.path.exists(file_path):
        return send_file(file_path)

    return "Excel file not found"


#############################################
# DELETE PROJECT
#############################################


@app.route("/delete_project/<code>")
def delete_project(code):

    if session.get("role") != "admin":
        return "Access Denied"

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM projects WHERE project_code=%s", (code,))
    conn.commit()

    file_path = os.path.join(EXCEL_FOLDER, code + ".xlsx")

    if os.path.exists(file_path):
        os.remove(file_path)

    cursor.close()
    conn.close()

    return redirect("/dashboard")


#############################################
# ID GENERATOR PAGE
#############################################


@app.route("/id_generator")
def id_generator():

    if "admin" not in session:
        return redirect("/")

    conn = connect_db()
    cursor = conn.cursor()

    cursor.execute(
        """
        SELECT id,
               first_name,
               last_name,
               position,
               photo,
               project_id
        FROM employees
        """
    )

    employees = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template("id_generator.html", employees=employees)


#############################################
# GENERATE ID
#############################################


@app.route("/generate_id/<code>/<employee_id>", methods=["GET", "POST"])
def generate_id(code, employee_id):

    conn = connect_db()
    cursor = conn.cursor()

    #################################
    # GET EMPLOYEE
    #################################

    cursor.execute(
        """
        SELECT first_name,last_name,position,photo
        FROM employees
        WHERE id=%s
        """,
        (employee_id,),
    )

    emp = cursor.fetchone()

    if not emp:
        cursor.close()
        conn.close()
        return "Employee not found"

    name = emp[0] + " " + emp[1]
    position = emp[2]
    photo_url = emp[3]

    photo = Image.open(photo_url)

    #################################
    # IF FORM SUBMITTED
    #################################

    if request.method == "POST":

        id_number = request.form["id_number"]
        expiry = request.form["expiry"]
        address = request.form.get("address", "")
        contact_number = request.form.get("contact_number", "")

        #################################
        # LOAD ID TEMPLATES (PIXEL PERFECT)
        #################################

        front = Image.open("id_templates/front.png").convert("RGB")
        back = Image.open("id_templates/back.png").convert("RGB")

        front = front.resize((600, 900))
        back = back.resize((600, 900))

        draw_front = ImageDraw.Draw(front)
        draw_back = ImageDraw.Draw(back)

        #################################
        # LOAD PROFESSIONAL FONTS
        #################################

        try:
            font_big = ImageFont.truetype("arialbd.ttf", 42)
            font_small = ImageFont.truetype("arial.ttf", 26)
        except:
            font_big = ImageFont.load_default()
            font_small = ImageFont.load_default()

        #################################
        # FIX 2X2 PHOTO PERFECTLY
        #################################

        size = min(photo.size)

        left = (photo.width - size) // 2
        top = (photo.height - size) // 2
        right = left + size
        bottom = top + size

        photo = photo.crop((left, top, right, bottom))
        photo = photo.resize((200, 200))

        #################################
        # EXACT PHOTO BOX LOCATION
        #################################

        PHOTO_X = 200
        PHOTO_Y = 210

        front.paste(photo, (PHOTO_X, PHOTO_Y))

        #################################
        # CENTER TEXT FUNCTION
        #################################

        def center_text(draw, text, font, y, width=600):

            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            x = (width - text_width) // 2
            draw.text((x, y), text, (0, 0, 0), font)

        #################################
        # FRONT TEXT (PIXEL PERFECT)
        #################################

        center_text(draw_front, name, font_big, 450)
        center_text(draw_front, position, font_small, 520)
        center_text(draw_front, "ID No: " + id_number, font_small, 570)

        #################################
        # BACK TEXT
        #################################

        draw_back.text((180, 320), name, (0, 0, 0), font_small)
        draw_back.text((180, 360), address, (0, 0, 0), font_small)
        draw_back.text((180, 400), contact_number, (0, 0, 0), font_small)
        draw_back.text((220, 740), "EXPIRY: " + expiry, (0, 0, 0), font_small)

        #################################
        # SAVE ID CARDS
        #################################

        if not os.path.exists("id_cards"):
            os.makedirs("id_cards")

        front_file = "id_cards/" + id_number + "_front.png"
        back_file = "id_cards/" + id_number + "_back.png"

        front.save(front_file, quality=100)
        back.save(back_file, quality=100)

        #################################
        # GET PROJECT CODE
        #################################

        cursor.execute(
            """
            SELECT project_code
            FROM projects
            WHERE id=%s
            """,
            (code,),
        )

        project = cursor.fetchone()

        if not project:
            cursor.close()
            conn.close()
            return "Project not found"

        project_code = project[0]

        #################################
        # OPEN EXCEL
        #################################

        excel_path = "excel_files/" + project_code + ".xlsx"
        wb = load_workbook(excel_path)

        #################################
        # ID SHEET
        #################################

        if "ID" not in wb.sheetnames:

            ws = wb.create_sheet("ID")
            ws.append(["NAME", "ID NUMBER", "EXPIRY", "IMAGE"])

        else:
            ws = wb["ID"]

        #################################
        # DUPLICATE CHECK
        #################################

        for r in ws.iter_rows(min_row=2):
            if r[1].value == id_number:
                return "ID Number Exists"

        #################################
        # SAVE TO EXCEL
        #################################

        row = ws.max_row + 2

        ws.cell(row=row, column=1).value = name
        ws.cell(row=row, column=2).value = id_number
        ws.cell(row=row, column=3).value = expiry

        img = ExcelImage(front_file)
        img.width = 250
        img.height = 400

        ws.row_dimensions[row].height = 300

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 70

        ws.add_image(img, "D" + str(row))

        wb.save(excel_path)

        #################################
        # REDIRECT
        #################################

        return redirect("/print_id/" + id_number)

    #################################
    # SHOW PAGE
    #################################

    cursor.close()
    conn.close()

    return render_template(
        "generate_id.html",
        emp=emp,
        code=code,
        employee_id=employee_id,
    )


#############################################
# PRINT PAGE
#############################################


@app.route("/print_id/<id_number>")
def print_id(id_number):

    front_file = "id_cards/" + id_number + "_front.png"
    back_file = "id_cards/" + id_number + "_back.png"

    return render_template("print_id.html", front=front_file, back=back_file)


#############################################
# SERVE IMAGES
#############################################


@app.route("/id_cards/<filename>")
def id_cards(filename):

    return send_from_directory("id_cards", filename)


#############################################
# SERVE UPLOADED FILES
#############################################


@app.route("/uploads/<path:filename>")
def uploaded_files(filename):

    return send_from_directory("uploads", filename)


#############################################
# RESET SYSTEM WITH PASSWORD CONFIRMATION
#############################################


@app.route("/reset_system", methods=["GET", "POST"])
def reset_system():

    if "admin" not in session:
        return redirect("/")

    if request.method == "POST":

        password = request.form["password"]

        conn = connect_db()
        cursor = conn.cursor()

        cursor.execute(
            "SELECT password FROM admins WHERE username=%s", (session["admin"],)
        )

        admin = cursor.fetchone()

        if not admin or not check_password_hash(admin[0], password):
            cursor.close()
            conn.close()
            return "Incorrect password. Reset cancelled."

        #################################
        # RESET DATABASE
        #################################

        cursor.execute("TRUNCATE TABLE employees RESTART IDENTITY CASCADE")
        cursor.execute("TRUNCATE TABLE projects RESTART IDENTITY CASCADE")

        conn.commit()

        cursor.close()
        conn.close()

        #################################
        # DELETE ID CARDS
        #################################

        if os.path.exists("id_cards"):
            for file in os.listdir("id_cards"):
                os.remove(os.path.join("id_cards", file))

        #################################
        # DELETE UPLOAD FILES
        #################################

        upload_folders = [
            "uploads/photos",
            "uploads/nbi",
            "uploads/certificates",
            "uploads/signatures",
            "uploads/secid",
            "uploads/wah",
        ]

        for folder in upload_folders:

            if os.path.exists(folder):

                for file in os.listdir(folder):

                    path = os.path.join(folder, file)

                    if os.path.isfile(path):
                        os.remove(path)

        #################################
        # DELETE PROJECT EXCEL FILES
        #################################

        if os.path.exists("excel_files"):

            for file in os.listdir("excel_files"):

                path = os.path.join("excel_files", file)

                if os.path.isdir(path):
                    continue

                if file.endswith(".xlsx"):
                    os.remove(path)

        return redirect("/dashboard")

    return render_template("reset_confirm.html")


#############################################
# LOGOUT
#############################################


@app.route("/logout")
def logout():

    session.clear()
    return redirect("/")


#############################################
# RUN SERVER
#############################################

if __name__ == "__main__":
    app.run(debug=True)
